"""
Excel loader for JP Morgan Asset Management fund data.
Implements ETL loading pattern with professional Excel formatting.
"""

import os
import shutil
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from ..models.fund_models import FundModel, DailyReport


class ExcelLoader:
    """Loader for saving fund data to professionally formatted Excel files."""
    
    def __init__(self, output_dir: str = "data"):
        """
        Initialize Excel loader.
        
        Args:
            output_dir: Directory for output files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Create subdirectories
        (self.output_dir / "current").mkdir(exist_ok=True)
        (self.output_dir / "backups").mkdir(exist_ok=True)
        (self.output_dir / "reports").mkdir(exist_ok=True)
        
        logger.info(f"Excel loader initialized with output directory: {self.output_dir}")
    
    def save_to_excel(self, funds: List[FundModel], filename: Optional[str] = None) -> str:
        """
        Save fund data to Excel with professional formatting.
        
        Args:
            funds: List of FundModel objects
            filename: Optional custom filename
            
        Returns:
            Path to saved Excel file
        """
        if not funds:
            raise ValueError("No fund data to save")
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"jp_morgan_funds_{timestamp}.xlsx"
        
        filepath = self.output_dir / "current" / filename
        
        # Convert to DataFrame
        df = self._funds_to_dataframe(funds)
        
        # Save with formatting
        self._save_formatted_excel(df, filepath)
        
        logger.success(f"Saved {len(funds)} funds to Excel: {filepath}")
        return str(filepath)
    
    def _funds_to_dataframe(self, funds: List[FundModel]) -> pd.DataFrame:
        """Convert list of FundModel objects to pandas DataFrame."""
        data = []
        
        for fund in funds:
            row = {
                'Ticker': fund.ticker,
                'CUSIP': fund.cusip,
                'Fund Name': fund.fund_name or '',
                '30 Day SEC Yield (%)': float(fund.sec_yield_30_day) if fund.sec_yield_30_day is not None else None,
                '30 Day SEC Yield Unsubsidized (%)': float(fund.sec_yield_30_day_unsubsidized) if fund.sec_yield_30_day_unsubsidized is not None else None,
                'Source URL': fund.url,
                'Scraped At': fund.scraped_at
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Sort by ticker for consistency
        df = df.sort_values('Ticker').reset_index(drop=True)
        
        logger.debug(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
        return df
    
    def _save_formatted_excel(self, df: pd.DataFrame, filepath: Path):
        """Save DataFrame to Excel with professional formatting."""
        # Save basic Excel file first
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Fund Data', index=False)
        
        # Apply professional formatting
        self._apply_excel_formatting(filepath)
    
    def _apply_excel_formatting(self, filepath: Path):
        """Apply professional formatting to Excel file."""
        workbook = load_workbook(filepath)
        worksheet = workbook['Fund Data']
        
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        border = Border(
            left=Side(border_style='thin', color='C0C0C0'),
            right=Side(border_style='thin', color='C0C0C0'),
            top=Side(border_style='thin', color='C0C0C0'),
            bottom=Side(border_style='thin', color='C0C0C0')
        )
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Apply data formatting
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # Auto-adjust column widths
        self._adjust_column_widths(worksheet)
        
        # Apply number formatting for percentage columns
        percentage_columns = ['D', 'E']  # 30 Day SEC Yield columns
        for col in percentage_columns:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '0.00%'
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
        
        # Add title and metadata
        self._add_title_and_metadata(worksheet)
        
        workbook.save(filepath)
        logger.debug(f"Applied professional formatting to {filepath}")
    
    def _adjust_column_widths(self, worksheet):
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_title_and_metadata(self, worksheet):
        """Add title and metadata to the worksheet."""
        # Insert rows at the top for title and metadata
        worksheet.insert_rows(1, 3)
        
        # Add title
        title_cell = worksheet['A1']
        title_cell.value = 'JP Morgan Asset Management Fund Data'
        title_cell.font = Font(name='Calibri', size=14, bold=True, color='366092')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Add generation timestamp
        timestamp_cell = worksheet['A2']
        timestamp_cell.value = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        timestamp_cell.font = Font(name='Calibri', size=10, color='666666')
        
        # Merge title cell across columns
        worksheet.merge_cells('A1:G1')
        worksheet.merge_cells('A2:G2')
    
    def create_backup(self, source_file: str) -> Optional[str]:
        """
        Create backup of existing Excel file.
        
        Args:
            source_file: Path to source file
            
        Returns:
            Path to backup file or None if source doesn't exist
        """
        source_path = Path(source_file)
        if not source_path.exists():
            logger.warning(f"Source file not found for backup: {source_file}")
            return None
        
        # Generate backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{source_path.stem}_backup_{timestamp}{source_path.suffix}"
        backup_path = self.output_dir / "backups" / backup_filename
        
        # Copy file
        shutil.copy2(source_path, backup_path)
        
        logger.info(f"Created backup: {backup_path}")
        return str(backup_path)
    
    def compare_with_previous(self, current_funds: List[FundModel], previous_file: Optional[str] = None) -> Dict[str, Any]:
        """
        Compare current funds with previous day's data.
        
        Args:
            current_funds: Current fund data
            previous_file: Path to previous Excel file
            
        Returns:
            Comparison report dictionary
        """
        if not previous_file or not Path(previous_file).exists():
            logger.info("No previous file found for comparison")
            return {
                'new_funds': len(current_funds),
                'updated_funds': 0,
                'unchanged_funds': 0,
                'removed_funds': 0,
                'changes': []
            }
        
        # Load previous data
        try:
            previous_df = pd.read_excel(previous_file, sheet_name='Fund Data', header=3)  # Skip title rows
            previous_funds = self._dataframe_to_funds(previous_df)
        except Exception as e:
            logger.error(f"Failed to load previous data: {e}")
            return {'error': f"Could not load previous data: {e}"}
        
        # Create lookup dictionaries
        current_lookup = {f.ticker: f for f in current_funds}
        previous_lookup = {f.ticker: f for f in previous_funds}
        
        # Analyze changes
        new_funds = []
        updated_funds = []
        unchanged_funds = []
        removed_funds = []
        changes = []
        
        # Check for new and updated funds
        for ticker, fund in current_lookup.items():
            if ticker not in previous_lookup:
                new_funds.append(ticker)
            else:
                prev_fund = previous_lookup[ticker]
                fund_changes = self._compare_funds(fund, prev_fund)
                if fund_changes:
                    updated_funds.append(ticker)
                    changes.extend(fund_changes)
                else:
                    unchanged_funds.append(ticker)
        
        # Check for removed funds
        for ticker in previous_lookup:
            if ticker not in current_lookup:
                removed_funds.append(ticker)
        
        comparison_report = {
            'new_funds': len(new_funds),
            'updated_funds': len(updated_funds),
            'unchanged_funds': len(unchanged_funds),
            'removed_funds': len(removed_funds),
            'changes': changes,
            'new_fund_tickers': new_funds,
            'updated_fund_tickers': updated_funds,
            'removed_fund_tickers': removed_funds
        }
        
        logger.info(f"Comparison complete: {comparison_report}")
        return comparison_report
    
    def _dataframe_to_funds(self, df: pd.DataFrame) -> List[FundModel]:
        """Convert DataFrame back to FundModel objects."""
        funds = []
        
        for _, row in df.iterrows():
            try:
                fund_data = {
                    'ticker': row.get('Ticker', ''),
                    'cusip': row.get('CUSIP', ''),
                    'fund_name': row.get('Fund Name', ''),
                    'sec_yield_30_day': row.get('30 Day SEC Yield (%)', None),
                    'sec_yield_30_day_unsubsidized': row.get('30 Day SEC Yield Unsubsidized (%)', None),
                    'url': row.get('Source URL', ''),
                    'scraped_at': str(row.get('Scraped At', ''))
                }
                
                # Handle NaN values
                for key, value in fund_data.items():
                    if pd.isna(value):
                        fund_data[key] = None if key in ['sec_yield_30_day', 'sec_yield_30_day_unsubsidized'] else ''
                
                fund = FundModel(**fund_data)
                funds.append(fund)
                
            except Exception as e:
                logger.warning(f"Failed to convert row to FundModel: {e}")
                continue
        
        return funds
    
    def _compare_funds(self, current: FundModel, previous: FundModel) -> List[Dict[str, Any]]:
        """Compare two FundModel objects and return list of changes."""
        changes = []
        
        # Compare yields
        if current.sec_yield_30_day != previous.sec_yield_30_day:
            changes.append({
                'ticker': current.ticker,
                'field': '30 Day SEC Yield',
                'old_value': float(previous.sec_yield_30_day) if previous.sec_yield_30_day else None,
                'new_value': float(current.sec_yield_30_day) if current.sec_yield_30_day else None
            })
        
        if current.sec_yield_30_day_unsubsidized != previous.sec_yield_30_day_unsubsidized:
            changes.append({
                'ticker': current.ticker,
                'field': '30 Day SEC Yield Unsubsidized',
                'old_value': float(previous.sec_yield_30_day_unsubsidized) if previous.sec_yield_30_day_unsubsidized else None,
                'new_value': float(current.sec_yield_30_day_unsubsidized) if current.sec_yield_30_day_unsubsidized else None
            })
        
        # Compare other key fields
        if current.cusip != previous.cusip:
            changes.append({
                'ticker': current.ticker,
                'field': 'CUSIP',
                'old_value': previous.cusip,
                'new_value': current.cusip
            })
        
        if current.fund_name != previous.fund_name:
            changes.append({
                'ticker': current.ticker,
                'field': 'Fund Name',
                'old_value': previous.fund_name,
                'new_value': current.fund_name
            })
        
        return changes
    
    def generate_daily_report(self, report_data: DailyReport) -> str:
        """
        Generate daily summary report in Excel format.
        
        Args:
            report_data: DailyReport model with summary statistics
            
        Returns:
            Path to generated report file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"daily_report_{timestamp}.xlsx"
        report_path = self.output_dir / "reports" / report_filename
        
        # Create summary data
        summary_data = {
            'Metric': [
                'Total Funds Processed',
                'Successful Scrapes',
                'Failed Scrapes',
                'Success Rate (%)',
                'New Funds',
                'Updated Funds',
                'Unchanged Funds',
                'Execution Time (seconds)',
                'Data Quality Score'
            ],
            'Value': [
                report_data.total_funds,
                report_data.successful_scrapes,
                report_data.failed_scrapes,
                round((report_data.successful_scrapes / report_data.total_funds) * 100, 1) if report_data.total_funds > 0 else 0,
                report_data.new_funds,
                report_data.updated_funds,
                report_data.unchanged_funds,
                round(report_data.execution_time, 2),
                'N/A'  # Could be calculated if data quality analysis is available
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        
        # Save report
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Add error breakdown if available
            if report_data.errors:
                errors_data = {
                    'Error Type': list(report_data.errors.keys()),
                    'Count': list(report_data.errors.values())
                }
                errors_df = pd.DataFrame(errors_data)
                errors_df.to_excel(writer, sheet_name='Errors', index=False)
        
        # Apply formatting to report
        self._format_report(report_path)
        
        logger.success(f"Generated daily report: {report_path}")
        return str(report_path)
    
    def _format_report(self, filepath: Path):
        """Apply formatting to daily report Excel file."""
        workbook = load_workbook(filepath)
        
        # Format Summary sheet
        if 'Summary' in workbook.sheetnames:
            worksheet = workbook['Summary']
            self._apply_basic_formatting(worksheet)
        
        # Format Errors sheet if it exists
        if 'Errors' in workbook.sheetnames:
            worksheet = workbook['Errors']
            self._apply_basic_formatting(worksheet)
        
        workbook.save(filepath)
        logger.debug(f"Applied formatting to report: {filepath}")
    
    def _apply_basic_formatting(self, worksheet):
        """Apply basic formatting to a worksheet."""
        # Header formatting
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Data formatting
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        border = Border(
            left=Side(border_style='thin', color='C0C0C0'),
            right=Side(border_style='thin', color='C0C0C0'),
            top=Side(border_style='thin', color='C0C0C0'),
            bottom=Side(border_style='thin', color='C0C0C0')
        )
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Apply data formatting
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # Auto-adjust column widths
        self._adjust_column_widths(worksheet) 